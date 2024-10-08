import sys
import os
import pandas as pd
import openpyxl
import glob
import time
from geopy.distance import geodesic
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QLineEdit, QPushButton, QVBoxLayout, QWidget, QTextEdit, QMessageBox, QProgressBar, QHBoxLayout
from PyQt5.QtGui import QPixmap

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

class RCRAutomatorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('RCR Sheet Automator')
        self.setGeometry(100, 100, 800, 600)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        layout = QVBoxLayout()

        logo_path = resource_path('Boeing_full_logo.png')
        self.logo_label = QLabel(self)
        self.logo_label.setPixmap(QPixmap(logo_path).scaled(200, 100))
        layout.addWidget(self.logo_label)

        # Directory Path Input
        self.input1_label = QLabel("Enter the directory path where the files are stored:")
        layout.addWidget(self.input1_label)
        self.input1 = QLineEdit(self)
        layout.addWidget(self.input1)

        # Permanent File Name Input
        self.input2_label = QLabel("Enter the name of the permanent Excel file:")
        layout.addWidget(self.input2_label)
        self.input2 = QLineEdit(self)
        layout.addWidget(self.input2)

        # Monthly File Pattern Input with Example
        self.input3_label = QLabel("Enter the pattern for the monthly files (e.g., tbl_RCR_Candidate* if raw data files are called tbl_RCR_Candidate_List 2024-XX-XX.xlsx):")
        layout.addWidget(self.input3_label)
        self.input3 = QLineEdit(self)
        layout.addWidget(self.input3)

        # Sheet Name Input
        self.input4_label = QLabel("Enter the name of the sheet in the permanent file where the data should be added:")
        layout.addWidget(self.input4_label)
        self.input4 = QLineEdit(self)
        layout.addWidget(self.input4)

        # Starting Row Number Input
        self.input5_label = QLabel("Enter the starting row number:")
        layout.addWidget(self.input5_label)
        self.input5 = QLineEdit(self)
        layout.addWidget(self.input5)

        self.submit_button = QPushButton("Submit", self)
        self.submit_button.clicked.connect(self.onSubmit)
        layout.addWidget(self.submit_button)

        self.progress_bar = QProgressBar(self)
        layout.addWidget(self.progress_bar)

        self.status_text = QTextEdit(self)
        self.status_text.setReadOnly(True)
        layout.addWidget(self.status_text)

        central_widget.setLayout(layout)

    def update_progress(self, progress, status):
        self.progress_bar.setValue(progress)
        self.status_text.append(status)

    def onSubmit(self):
        input1 = self.input1.text()
        input2 = self.input2.text()
        input3 = self.input3.text()
        input4 = self.input4.text()
        input5 = self.input5.text()

        if input1 and input2 and input3 and input4 and input5:
            try:
                directory_path = input1
                permanent_file_name = input2
                monthly_file_pattern = input3
                sheet_name = input4
                start_row = int(input5)
                header_row = 3

                self.update_progress(10, "Looking for monthly files...")
                monthly_files = glob.glob(os.path.join(directory_path, monthly_file_pattern))

                if not monthly_files:
                    QMessageBox.critical(self, "Error", "No monthly files found matching the pattern.")
                    return

                latest_file = max(monthly_files, key=os.path.getctime)

                self.update_progress(20, f"Reading data from {latest_file}...")
                try:
                    monthly_data = pd.read_excel(latest_file, engine='openpyxl')
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Error reading the monthly file: {e}")
                    return

                self.update_progress(30, "Opening permanent file...")
                try:
                    permanent_wb = openpyxl.load_workbook(os.path.join(directory_path, permanent_file_name), keep_vba=True, read_only=False, data_only=False)
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Error opening the permanent file: {e}")
                    return

                if sheet_name not in permanent_wb.sheetnames:
                    QMessageBox.critical(self, "Error", f"Sheet '{sheet_name}' does not exist in the permanent file.")
                    return

                sheet = permanent_wb[sheet_name]
                self.update_progress(40, "Processing headers...")

                permanent_headers = [cell.value for cell in sheet[header_row] if cell.value is not None]
                monthly_headers = [header.strip().lower() for header in monthly_data.columns]

                self.update_progress(50, "Preparing data for update...")

                for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        if cell.data_type != 'f':
                            cell.value = None

                for col_idx, header in enumerate(permanent_headers):
                    normalized_header = header.strip().lower()
                    if normalized_header in monthly_headers:
                        monthly_col_idx = monthly_headers.index(normalized_header)
                        for row_idx, value in enumerate(monthly_data.iloc[:, monthly_col_idx]):
                            cell = sheet.cell(row=row_idx + start_row, column=col_idx + 1)
                            cell.value = value
                        if normalized_header == "on rcr events report":
                            break

                self.update_progress(60, "Updating data...")

                excel_file = pd.ExcelFile(os.path.join(directory_path, permanent_file_name))
                candidate_df = excel_file.parse('Top 30', header=2)
                facilitator_df = excel_file.parse('Facilitator List With Location')

                # Ensure 'RCR_Limit' column exists
                if 'RCR_Limit' not in facilitator_df.columns:
                    QMessageBox.critical(self, "Error", "The 'RCR_Limit' column is missing in the facilitator list.")
                    return

                candidate_df['Closest Facilitator'] = ''
                geolocator = Nominatim(user_agent='my_app')

                def geocode_with_retry(location):
                    max_retries = 5
                    retries = 0
                    while retries < max_retries:
                        try:
                            return geolocator.geocode(location)
                        except GeocoderTimedOut:
                            retries += 1
                            time.sleep(1)
                    return None

                facilitator_locations = {}
                facilitator_count = {}
                facilitator_df = facilitator_df.dropna()
                for facilitator_index, facilitator_row in facilitator_df.iterrows():
                    facilitator_location = facilitator_row['Location']
                    facilitator_name = facilitator_row['Name']
                    facilitator_state, facilitator_city = facilitator_location.split('-')
                    facilitator_location = f"{facilitator_city.strip()}, {facilitator_state.strip()}"
                    facilitator_coordinates = geocode_with_retry(facilitator_location)
                    facilitator_locations[facilitator_name] = facilitator_coordinates
                    facilitator_count[facilitator_name] = 0

                for index, candidate_row in candidate_df.head(30).iterrows():
                    candidate_location = f"{candidate_row['City']}, {candidate_row['State']}"
                    candidate_coordinates = geocode_with_retry(candidate_location)
                    if candidate_coordinates is None:
                        continue
                    candidate_latitude = candidate_coordinates.latitude
                    candidate_longitude = candidate_coordinates.longitude
                    closest_facilitator = None
                    min_distance = float('inf')

                    for facilitator_index, facilitator_row in facilitator_df.iterrows():
                        facilitator_coordinates = facilitator_locations[facilitator_row['Name']]
                        if facilitator_coordinates is None:
                            continue
                        facilitator_latitude = facilitator_coordinates.latitude
                        facilitator_longitude = facilitator_coordinates.longitude

                        distance = geodesic((candidate_latitude, candidate_longitude), (facilitator_latitude, facilitator_longitude)).miles

                        if distance < min_distance and facilitator_count[facilitator_row['Name']] < facilitator_row['RCR_Limit']:
                            min_distance = distance
                            closest_facilitator = facilitator_row['Name']

                    if closest_facilitator is not None:
                        candidate_df.at[index, 'Closest Facilitator'] = closest_facilitator
                        facilitator_count[closest_facilitator] = facilitator_count.get(closest_facilitator, 0) + 1
                    else:
                        print(f"No facilitator found for candidate at index {index}")

                ws = permanent_wb['Top 30']
                ws.cell(row=3, column=ws.max_column + 1, value='Closest Facilitator')

                for idx, value in enumerate(candidate_df['Closest Facilitator'], start=4):
                    ws.cell(row=idx, column=ws.max_column, value=value)

                permanent_wb.save(os.path.join(directory_path, permanent_file_name))

                self.update_progress(100, "Process completed successfully!")
                QMessageBox.information(self, "Success", f"Permanent file '{permanent_file_name}' updated with data from '{latest_file}' in sheet '{sheet_name}' starting at row {start_row}")
                QMessageBox.information(self, "Success", f"Updated 'Top 30' sheet in '{permanent_file_name}' with closest facilitator information.")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")
        else:
            QMessageBox.warning(self, "Warning", "Please fill in all fields before submitting.")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = RCRAutomatorApp()
    ex.show()
    sys.exit(app.exec_())
